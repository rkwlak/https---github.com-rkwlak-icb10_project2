import re
import datetime
import pandas as pd
import io
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 쇼핑 카테고리 매핑 (Naver 쇼핑 API에서 사용되는 카테고리 ID)
# ---------------------------------------------------------------------------
# HTML 리스트에 포함된 카테고리를 딕셔너리 형태로 관리합니다. UI에서 선택한
# 카테고리 ID를 API 파라미터에 전달하여 해당 카테고리 상품만 검색하도록 합니다.
SHOP_CATEGORY_MAP: dict[str, str] = {
    "패션의류": "50000000",
    "패션잡화": "50000001",
    "화장품/미용": "50000002",
    "디지털/가전": "50000003",
    "가구/인테리어": "50000004",
    "출산/육아": "50000005",
    "식품": "50000006",
    "스포츠/레저": "50000007",
    "생활/건강": "50000008",
    "여가/생활편의": "50000009",
    "면세점": "50000010",
    "도서": "50005542",
}



def parse_api_key(key_str: str):
    """클라이언트 ID와 시크릿을 "ID:SECRET" 형태로 입력받아 튜플 반환.
    입력이 비어 있거나 형식이 맞지 않으면 (None, None) 반환.
    """
    if not key_str:
        return None, None
    parts = key_str.strip().split(":")
    if len(parts) != 2:
        return None, None
    client_id, client_secret = parts[0].strip(), parts[1].strip()
    return client_id, client_secret


def split_keywords(raw: str):
    """쉼표(,) 혹은 줄바꿈으로 구분된 검색어 문자열을 리스트로 변환."""
    if not raw:
        return []
    # 콤마와 줄바꿈을 모두 구분자로 사용
    keywords = re.split(r"[\n,]+", raw)
    # 빈 문자열 제거 및 앞뒤 공백 정리
    return [kw.strip() for kw in keywords if kw.strip()]


def enforce_limit(keywords: list[str], max_count: int = 500) -> list[str]:
    """키워드 개수가 제한을 초과하면 앞쪽 `max_count` 개만 남긴다."""
    return keywords[:max_count]


def format_date(dt: datetime.date) -> str:
    """날짜 객체를 Naver API에서 요구하는 YYYY-MM-DD 문자열로 변환."""
    return dt.strftime("%Y-%m-%d")


def download_df_as_csv(df: pd.DataFrame) -> bytes:
    """DataFrame을 UTF‑8 CSV 바이너리로 변환해 Streamlit 다운로드에 사용."""
    return df.to_csv(index=False).encode("utf-8")


def get_word_frequencies(titles: list[str]) -> pd.DataFrame:
    """텍스트 리스트(예: 제목)에서 태그와 특수문자를 제거하고,
    간단한 한글 불용어를 필터링하여 상위 10개 단어 빈도 분석 DataFrame을 반환합니다.
    """
    words = []
    stopwords = {
        '은', '는', '이', '가', '을', '를', '의', '에', '와', '과', '으로', '로', '에서', '하고',
        '그리고', '네이버', '검색', '블로그', '카페', '뉴스', '및', '등', '더', '한', '하는', '할',
        '있는', '없어', '대해', '대한', '합니다', '한다', '있습니다', '최근', '추천', '비교', '후기'
    }
    for title in titles:
        if not isinstance(title, str):
            continue
        # HTML 태그 제거 (예: <b> 태그 등)
        clean_title = re.sub(r'<[^>]+>', '', title)
        # 특수문자를 제외한 단어 분리 (한글, 영어, 숫자 조합)
        tokens = re.findall(r'[가-힣a-zA-Z0-9]{2,15}', clean_title) # 2글자 이상 15글자 이하
        for token in tokens:
            # 숫자로만 이루어진 단어 제외 (예: 날짜, 순번 등)
            if token.isdigit():
                continue
            if token.lower() not in stopwords:
                words.append(token)
                
    if not words:
        return pd.DataFrame(columns=['단어', '빈도'])
        
    df_words = pd.DataFrame(words, columns=['단어'])
    counts = df_words['단어'].value_counts().head(10).reset_index()
    counts.columns = ['단어', '빈도']
    return counts


def download_df_as_excel(df: pd.DataFrame) -> bytes:
    """DataFrame을 헤더 스타일링 및 열 너비 자동 맞춤이 적용된
    고급 엑셀 파일(.xlsx) 바이너리로 변환하여 반환합니다.
    """
    output = io.BytesIO()
    # openpyxl 엔진을 사용하여 ExcelWriter 생성
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='분석 데이터', index=False)
        
        # 스타일링을 위한 워크시트 가져오기
        workbook = writer.book
        worksheet = writer.sheets['분석 데이터']
        
        # 헤더 채우기 스타일 (남색 배경에 굵은 흰색 글씨, 가운데 정렬)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # 데이터 셀 스타일 (기본 글꼴 설정)
        data_font = Font(name='맑은 고딕', size=10)
        data_alignment = Alignment(vertical='center')
        
        # 헤더 스타일 설정
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        # 데이터 셀 글꼴 및 자동 줄바꿈 설정, 열 너비 자동 맞춤 계산
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # 데이터 행에 스타일 적용
                if cell.row > 1:
                    cell.font = data_font
                    cell.alignment = data_alignment
                
                # 셀 값의 길이를 계산하여 최댓값 추적
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            
            # 한글 인코딩 등을 감안하여 여유 폭 지정 (최소 12, 최대 50)
            adjusted_width = max(max_len * 1.5, 12)
            adjusted_width = min(adjusted_width, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width
            
        # 격자선이 보이도록 설정
        worksheet.views.sheetView[0].showGridLines = True
        
    return output.getvalue()
